from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from user_foods_utils import UserFood, parse_food_nutrients
from utils import generate_random_filename
from user_foods_editing_utils import send_user_foods_file, edit_download_processed_count_message_text


def get_user_foods_from_excel_file(file_path) -> List[UserFood]:
    try:
        main_rows, user_rows, user_column_titles = _parse_excel_table(file_path)
        user_foods:List[UserFood] = []
        full_nutrients_str = None
        for main_row_index, row in enumerate(main_rows):
            if user_rows and user_column_titles:
                full_nutrients = {}
                for i, user_data_value in enumerate(user_rows[main_row_index]):
                    data_title = user_column_titles[i]
                    full_nutrients[data_title] = user_data_value

                full_nutrients_str = ','.join(f'{title}:{value}' for title, value in full_nutrients.items())
                
            user_foods.append(UserFood(id=None,
                                user_id=None,
                                barcode=row[1],
                                title=row[0],
                                group_name=row[2],
                                weight=row[3],
                                one_piece_weight=row[4],
                                calories=row[5],
                                fat=row[6],
                                protein=row[7],
                                carbohydrate=row[8],
                                full_nutrients=full_nutrients_str,
                                is_animal_protein=row[9],
                                is_unsaturated_fats=row[10],
                                is_complex_carbohydrates=row[11]))
            
        return user_foods
    
    except Exception as e:
        print('Ошибка! get_user_foods_from_excel_file()', e)
        return None

def _parse_excel_table(file_path):
    parsed_rows = []
    other_data = []

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        # Берем первые 12 столбцов
        parsed_row = list(row[:12])
        parsed_rows.append(parsed_row)

        # Обрабатываем оставшиеся столбцы и добавляем в отдельный список
        other_columns = list(row[12:])
        other_data.append(other_columns)

    return parsed_rows[1:], other_data[1:], list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0][12:]

def _get_column_index(sheet, column_name):
    column_index = None

    # Проверяем существование колонки с заданным названием
    for col_idx, column in enumerate(sheet.iter_cols(), start=1):
        if column[0].value == column_name:
            column_index = col_idx
            break

    return column_index

def _get_full_nutrients_titles(user_foods:List[UserFood]):
    full_nutrients_titles = []
    for food in user_foods:
        full_nutrients = parse_food_nutrients(food.full_nutrients)
        # Получаем список из названий ключей
        full_nutrients_titles += list(full_nutrients.keys())

    return list(set(full_nutrients_titles))

def get_user_foods_table(user_foods:List[UserFood], message):
    user_foods_count = len(user_foods)
    try:
        # Создаем новую таблицу
        wb = Workbook()
        ws = wb.active

        # Заполняем заголовки таблицы
        headers = ["ID", "Название", "Штрихкод", "Категория", "Вес (грамм)", "Вес (грамм) в 1 шт.",
                "Калории", "Белки (грамм)", "Жиры (грамм)", "Углеводы (грамм)",
                "Животный белок?", "Ненасыщенные жиры?", "Сложные углеводы?"]
        full_nutrients_titles = sorted(_get_full_nutrients_titles(user_foods))
        ws.append(headers + full_nutrients_titles)

        # Заполняем таблицу данными из экземпляров класса UserFood
        for processed_count, food_item in enumerate(user_foods):
            processed_count += 1
            data_row = [food_item.id, food_item.title, food_item.barcode,
                        food_item.group_name, food_item.weight, food_item.one_piece_weight,
                        food_item.calories, food_item.protein, food_item.fat, food_item.carbohydrate,
                        food_item.is_animal_protein, food_item.is_unsaturated_fats,
                        food_item.is_complex_carbohydrates]
            ws.append(data_row)
            # Вставляем значения из словаря full_nutrients в таблицу
            full_nutrients = full_nutrients = parse_food_nutrients(food_item.full_nutrients)
            for title, value in full_nutrients.items():
                column_index = _get_column_index(ws, title)
                ws.cell(row=ws.max_row, column=column_index, value=value)


            edit_download_processed_count_message_text(message, user_foods_count, processed_count)

        # Создаем объекты PatternFill для светлозеленого и желтого цветов
        light_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # Применяем цвета заливки к ячейкам в первой строке
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = yellow_fill if col_idx <= 13 else light_green_fill
        # Создаем объект Border для настройки стиля границ
        thin_border = Border(left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"))
        # Применяем границы ко всем ячейкам
        for row in ws.iter_rows(min_row=1, min_col=1):
            for cell in row:
                cell.border = thin_border

        filename = f'{generate_random_filename()}.xlsx'
        filepath = f'user_foods_tables/{filename}'
        wb.save(filepath)
        edit_download_processed_count_message_text(message, is_finished=True)

    except Exception as e:
        print('get_user_foods_table()', e)
        filepath = None

    finally:
        send_user_foods_file(message, filepath)