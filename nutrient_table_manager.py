from openpyxl import Workbook
from openpyxl.styles import PatternFill
from nutritionix_utils import NutritionixNutrient
from typing import List
from os import remove
from utils import generate_random_filename, get_now_date
from openpyxl.utils import get_column_letter

def generate_table(nutrients:List[NutritionixNutrient]):
    try:
        # Создаем новую рабочую книгу и активный лист
        workbook = Workbook()
        sheet = workbook.active

        # Заголовок таблицы
        header = ["Название (англ.)", "Название (автоперевод)", "Количество", "Единица измерения"]
        sheet.append(header)

        # Заполняем таблицу данными из списка объектов
        for nutrient in nutrients:
            row = [nutrient.name, nutrient.ru_name, nutrient.value, nutrient.unit]
            sheet.append(row)

        # Расширяем ширину всех колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Получаем букву колонки
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Устанавливаем ширину с запасом
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Создаем стиль для заливки ячеек заголовка
        header_fill = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")

        # Заливаем ячейки заголовка
        for cell in sheet[1]:
            cell.fill = header_fill

        file_path = f'nutrient_tables/Отчет о нутриентах от {get_now_date("%d.%m.%Y")} ({generate_random_filename(6)}).xlsx'
        # Сохраняем рабочую книгу в файл
        workbook.save(file_path)
        return file_path
    except Exception as e:
        print('Ошибка в generate_table()', e)
        return None

def delete_table(file_path):
    try:
        remove(file_path)
    except Exception as e:
        print('Ошибка в delete_table()', e)